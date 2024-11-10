import os
import constants
from openpyxl import load_workbook

def remove_blank(list):
    """リストの各項目の末尾にある空白文字を削除
    """
    return [element.rstrip() if isinstance(element, str) else element for element in list]

def extract_file_name():
    """csvファイルのうち最新2つのファイル名を取得
    """
    
    # ファイル一覧の取得
    file_names = os.listdir('./dbdata')
    
    # 名前順に並び替え
    sorted_file_names = sorted(file_names)

    # 最新の2つのファイル名を返却
    return [sorted_file_names[-2], sorted_file_names[-1]]

def extract_unv_data(data_rows):
    """
    大学情報のデータのみを抽出する関数
    
    Args:
        data_rows (List[Tuple[str, str, str, str, str, str, str]]): 全情報の2次元配列
    
    Returns:
        Set{Tuple[str, str, str, str, str, str, str]}: 大学情報だけを含む新しいリスト
    """
    
    return {tup for tup in data_rows if int(tup[0][0]) < constants.BORDER_UNV_VOC}

def extract_voc_data(data_rows):
    """
    専門学校情報のデータのみを抽出する関数
    
    Args:
        data (List[Tuple[str, str, str, str, str, str, str]]): 全情報の2次元配列
    
    Returns:
        Set[Tuple[str, str, str, str, str, str, str]]: 専門学校情報だけを含む新しいリスト
    """
    
    return {tup for tup in data_rows if int(tup[0][0]) >= constants.BORDER_UNV_VOC}

def extract_dif_by_compare_cd(base, compare):
    """baseにはあるがcompareにはないものを比較用コードを基に抽出
    """

    # compareの比較用コードのみのsetを作成
    compare_compare_cd_set = {item[constants.ROW_COMPARE_CD] for item in compare}
    
    # baseにはあるがcompareにはない1列目の要素
    unique_in_base = {item for item in base if item[constants.ROW_COMPARE_CD] not in compare_compare_cd_set}
    
    return unique_in_base
    

def to_excel(data, sheet_name, overwrite):
    """エクセルを出力
    """
    
    if overwrite:
        wb = load_workbook('./result/結果.xlsx')
    else:
        wb = load_workbook('./result/フォーマット.xlsx')

    ws = wb[sheet_name]
    
    for row_index, row_data in enumerate(data):
        for col_index, col_data in enumerate(row_data):
            ws.cell(row=row_index + 2,
                       column=col_index + 1,
                       value=col_data)
    
    wb.save("./result/結果.xlsx")

if __name__ == '__main__':
    test = {
        ('10001010', 'テスト高校'),
        ('2000', 'テスト高校2'),
        ('7000', 'テスト高校3'),
        ('8000', 'テスト高校4'),
        ('4000', 'テスト高校2'),
    }
    
    test2 = {
        ('10001010', 'テスト高校'),
        ('2000', 'テスト高校2'),
        ('7000', 'テスト高校3'),
        ('8000', 'テスト高校4'),
        ('4000', 'テスト高校2'),
        ('9000', 'テスト高校2'),
        ('3000', 'テスト高校2'),
    }
    